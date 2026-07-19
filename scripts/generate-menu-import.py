"""Generate the restaurant/menu migration from the supplied Vantoo workbook."""

from __future__ import annotations

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl


RESTAURANTS = [
    {
        "sheet": "Kekiz Cake Shop Anvi Foods",
        "id": "r-kekiz-anvi",
        "name": "Kekiz Cake Shop Anvi Foods",
        "cuisine": ["Cakes", "Bakery", "Desserts"],
        "image": "https://images.unsplash.com/photo-1578985545062-69928b1d9587?w=900&q=80",
        "parser": "template",
    },
    {
        "sheet": "Aroma ",
        "id": "r-aroma",
        "name": "Aroma",
        "cuisine": ["Cafe", "Fast Food", "Beverages"],
        "image": "https://images.unsplash.com/photo-1554118811-1e0d58224f24?w=900&q=80",
        "parser": "aroma",
    },
    {
        "sheet": "Balaji A One Bhel",
        "id": "r-balaji-a-one-bhel",
        "name": "Balaji A One Bhel",
        "cuisine": ["Chaat", "Street Food", "Snacks"],
        "image": "https://images.unsplash.com/photo-1601050690597-df0568f70950?w=900&q=80",
        "parser": "template",
    },
    {
        "sheet": "Bhosle Kichan Pur Veg",
        "id": "r-bhosle-kitchen-pure-veg",
        "name": "Bhosle Kitchen Pure Veg",
        "cuisine": ["Pure Veg", "North Indian", "Chinese"],
        "image": "https://images.unsplash.com/photo-1585937421612-70a008356fbe?w=900&q=80",
        "parser": "template",
    },
    {
        "sheet": "Chai Vaai cafe",
        "id": "r-chai-vaai-cafe",
        "name": "Chai Vaai Cafe",
        "cuisine": ["Cafe", "Fast Food", "Beverages"],
        "image": "https://images.unsplash.com/photo-1495474472287-4d71bcdd2085?w=900&q=80",
        "parser": "simple",
    },
    {
        "sheet": "Datta Bhel",
        "id": "r-datta-bhel",
        "name": "Datta Bhel",
        "cuisine": ["Bhel", "Chaat", "Street Food"],
        "image": "https://images.unsplash.com/photo-1601050690117-94f5f6fa8bd7?w=900&q=80",
        "parser": "template",
    },
    {
        "sheet": "Happy Bites",
        "id": "r-happy-bites",
        "name": "Happy Bites",
        "cuisine": ["Desserts", "Waffles", "Beverages"],
        "image": "https://images.unsplash.com/photo-1562376552-0d160a2f238d?w=900&q=80",
        "parser": "template",
    },
]

GROCERY_IMAGE = (
    "https://images.unsplash.com/photo-1610348725531-843dff563e2c?w=900&q=80"
)


def slugify(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode()
    value = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return value or "item"


def sql(value: object) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def numeric_price(value: object) -> float | None:
    if isinstance(value, (int, float)) and value > 0:
        return float(value)
    return None


def template_rows(sheet, restaurant: dict) -> list[dict]:
    rows: list[dict] = []
    current_category = ""
    for row_number in range(2, sheet.max_row + 1):
        category = sheet.cell(row_number, 2).value
        if category not in (None, ""):
            current_category = str(category).strip()
        name = sheet.cell(row_number, 6).value
        raw_price = sheet.cell(row_number, 7).value
        if name in (None, ""):
            continue

        name = str(name).strip()
        split_prices = (
            re.fullmatch(r"\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)\s*", raw_price)
            if isinstance(raw_price, str)
            else None
        )
        if split_prices:
            for size, price in zip(("Half", "Full"), split_prices.groups()):
                rows.append(
                    make_product(
                        restaurant,
                        row_number,
                        current_category or "Menu",
                        f"{name} ({size})",
                        float(price),
                        None,
                        None,
                        size,
                        id_suffix=size.lower(),
                    )
                )
            continue

        price = numeric_price(raw_price)
        if price is None:
            continue
        original_price = numeric_price(sheet.cell(row_number, 15).value)
        if original_price is not None and original_price <= price:
            original_price = None
        description = sheet.cell(row_number, 8).value
        image = sheet.cell(row_number, 10).value
        rows.append(
            make_product(
                restaurant,
                row_number,
                current_category or "Menu",
                name,
                price,
                original_price,
                str(description).strip() if description not in (None, "") else None,
                category_unit(current_category),
                str(image).strip() if isinstance(image, str) and image.startswith("http") else None,
            )
        )
    return rows


def aroma_rows(sheet, restaurant: dict) -> list[dict]:
    rows = []
    for row_number in range(2, sheet.max_row + 1):
        name = sheet.cell(row_number, 3).value
        price = numeric_price(sheet.cell(row_number, 4).value)
        if name in (None, "") or price is None:
            continue
        rows.append(
            make_product(
                restaurant,
                row_number,
                "Cafe Menu",
                str(name).strip(),
                price,
                None,
                None,
                None,
            )
        )
    return rows


def simple_rows(sheet, restaurant: dict) -> list[dict]:
    rows = []
    for row_number in range(1, sheet.max_row + 1):
        category = sheet.cell(row_number, 1).value
        name = sheet.cell(row_number, 2).value
        price = numeric_price(sheet.cell(row_number, 3).value)
        if name in (None, "") or price is None:
            continue
        rows.append(
            make_product(
                restaurant,
                row_number,
                str(category).strip() if category not in (None, "") else "Menu",
                str(name).strip(),
                price,
                None,
                None,
                None,
            )
        )
    return rows


def category_unit(category: str) -> str | None:
    normalized = category.lower()
    if "350" in normalized:
        return "350 g"
    if "half kg" in normalized:
        return "500 g"
    if "1 kg" in normalized:
        return "1 kg"
    return None


def make_product(
    restaurant: dict,
    row_number: int,
    category: str,
    name: str,
    price: float,
    original_price: float | None,
    description: str | None,
    unit: str | None,
    image: str | None = None,
    id_suffix: str = "",
) -> dict:
    suffix = f"-{id_suffix}" if id_suffix else ""
    return {
        "id": f"menu-{slugify(restaurant['id'])}-{row_number}{suffix}",
        "name": name,
        "description": description or f"{category} from {restaurant['name']}.",
        "service": "food",
        "category": category,
        "brand": restaurant["name"],
        "price": price,
        "original_price": original_price,
        "rating": 0,
        "reviews": 0,
        "image": image or restaurant["image"],
        "images": [image or restaurant["image"]],
        "attributes": {
            "vegetarian": True,
            "source_sheet": restaurant["sheet"].strip(),
            "source_row": row_number,
        },
        "vendor_id": restaurant["id"],
        "unit": unit,
        "in_stock": True,
    }


def grocery_rows(sheet) -> list[dict]:
    rows = []
    seen = set()
    for row_number in range(1, sheet.max_row):
        name = sheet.cell(row_number, 2).value
        unit = sheet.cell(row_number + 1, 2).value
        raw_price = sheet.cell(row_number + 2, 2).value
        if not (
            isinstance(name, str)
            and isinstance(unit, str)
            and re.fullmatch(r"\d+(?:\.\d+)?\s*(?:g|kg|ml|l)", unit.strip(), re.I)
            and isinstance(raw_price, str)
            and re.search(r"\d", raw_price)
        ):
            continue
        price_match = re.search(r"(\d+(?:\.\d+)?)", raw_price)
        if not price_match:
            continue
        price = float(price_match.group(1))
        original_price = None
        next_value = sheet.cell(row_number + 3, 2).value
        if isinstance(next_value, str) and "₹" in next_value:
            original_match = re.search(r"(\d+(?:\.\d+)?)", next_value)
            if original_match:
                original_price = float(original_match.group(1))
        key = (name.casefold(), unit.casefold(), price)
        if key in seen:
            continue
        seen.add(key)
        rows.append(
            {
                "id": f"menu-fresh-vegetables-{slugify(name)}-{row_number}",
                "name": name.strip(),
                "description": "Fresh vegetable.",
                "service": "grocery",
                "category": "Vegetables",
                "brand": "Fresh Vegetables",
                "price": price,
                "original_price": original_price,
                "rating": 0,
                "reviews": 0,
                "image": GROCERY_IMAGE,
                "images": [GROCERY_IMAGE],
                "attributes": {
                    "source_sheet": "Sheet9",
                    "source_row": row_number,
                },
                "vendor_id": None,
                "unit": unit.strip(),
                "in_stock": True,
            }
        )
    return rows


def dedupe(products: list[dict]) -> list[dict]:
    result = []
    seen = set()
    for product in products:
        key = (
            product["vendor_id"],
            product["category"].casefold(),
            product["name"].casefold(),
            product["price"],
        )
        if key in seen:
            continue
        seen.add(key)
        result.append(product)
    return result


def generate(workbook_path: Path) -> str:
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    products: list[dict] = []
    for restaurant in RESTAURANTS:
        sheet = workbook[restaurant["sheet"]]
        parser = restaurant["parser"]
        if parser == "template":
            products.extend(template_rows(sheet, restaurant))
        elif parser == "aroma":
            products.extend(aroma_rows(sheet, restaurant))
        else:
            products.extend(simple_rows(sheet, restaurant))
    products.extend(grocery_rows(workbook["Sheet9"]))
    products = dedupe(products)
    product_ids = [product["id"] for product in products]
    if len(product_ids) != len(set(product_ids)):
        raise ValueError("Generated duplicate product IDs")

    lines = [
        "-- Restaurant and menu import generated from Vantoo menu.xlsx.",
        f"-- Imported {len(RESTAURANTS)} restaurants and {len(products)} priced products.",
        "begin;",
        "",
        "create table if not exists public.restaurants (",
        "  id text primary key,",
        "  name text not null,",
        "  cuisine text[] not null default '{}',",
        "  rating numeric not null default 0,",
        "  reviews integer not null default 0,",
        "  delivery_time text not null default '',",
        "  price_for_two numeric not null default 0,",
        "  image text not null default '',",
        "  offer text,",
        "  promoted boolean not null default false,",
        "  is_active boolean not null default true,",
        "  created_at timestamptz not null default now(),",
        "  updated_at timestamptz not null default now()",
        ");",
        "",
        "create index if not exists restaurants_active_idx",
        "  on public.restaurants (is_active, name);",
        "",
        "alter table public.restaurants enable row level security;",
        'drop policy if exists "Public read active restaurants" on public.restaurants;',
        'create policy "Public read active restaurants"',
        "  on public.restaurants for select to anon, authenticated",
        "  using (is_active = true);",
        "grant select on public.restaurants to anon, authenticated;",
        "",
        "insert into public.restaurants",
        "  (id, name, cuisine, rating, reviews, delivery_time, price_for_two, image, is_active)",
        "values",
    ]
    restaurant_values = []
    for restaurant in RESTAURANTS:
        cuisine = "array[" + ", ".join(sql(x) for x in restaurant["cuisine"]) + "]::text[]"
        restaurant_values.append(
            "  ("
            + ", ".join(
                [
                    sql(restaurant["id"]),
                    sql(restaurant["name"]),
                    cuisine,
                    "0",
                    "0",
                    "''",
                    "0",
                    sql(restaurant["image"]),
                    "true",
                ]
            )
            + ")"
        )
    lines.append(",\n".join(restaurant_values))
    lines.extend(
        [
            "on conflict (id) do update set",
            "  name = excluded.name,",
            "  cuisine = excluded.cuisine,",
            "  image = excluded.image,",
            "  is_active = excluded.is_active,",
            "  updated_at = now();",
            "",
            "insert into public.products",
            "  (id, name, description, service, category, brand, price, original_price,",
            "   rating, reviews, image, images, attributes, vendor_id, unit, in_stock, updated_at)",
            "values",
        ]
    )
    product_values = []
    for product in products:
        product_values.append(
            "  ("
            + ", ".join(
                [
                    sql(product["id"]),
                    sql(product["name"]),
                    sql(product["description"]),
                    sql(product["service"]),
                    sql(product["category"]),
                    sql(product["brand"]),
                    sql(product["price"]),
                    sql(product["original_price"]),
                    sql(product["rating"]),
                    sql(product["reviews"]),
                    sql(product["image"]),
                    sql(json.dumps(product["images"], ensure_ascii=False)) + "::jsonb",
                    sql(json.dumps(product["attributes"], ensure_ascii=False)) + "::jsonb",
                    sql(product["vendor_id"]),
                    sql(product["unit"]),
                    sql(product["in_stock"]),
                    "now()",
                ]
            )
            + ")"
        )
    lines.append(",\n".join(product_values))
    lines.extend(
        [
            "on conflict (id) do update set",
            "  name = excluded.name,",
            "  description = excluded.description,",
            "  service = excluded.service,",
            "  category = excluded.category,",
            "  brand = excluded.brand,",
            "  price = excluded.price,",
            "  original_price = excluded.original_price,",
            "  image = excluded.image,",
            "  images = excluded.images,",
            "  attributes = excluded.attributes,",
            "  vendor_id = excluded.vendor_id,",
            "  unit = excluded.unit,",
            "  in_stock = excluded.in_stock,",
            "  updated_at = now();",
            "",
            "commit;",
            "",
        ]
    )
    return "\n".join(lines)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit("Usage: generate-menu-import.py WORKBOOK OUTPUT_SQL")
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    output.write_text(generate(source), encoding="utf-8")
    print(f"Generated {output}")
